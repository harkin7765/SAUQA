import pydicom
from tkinter import filedialog
from tkinter import *
from PIL import Image
import cv2
from matplotlib import pyplot as plt
import os
import cmath
import ntpath
import copy
import openpyxl
import math
import numpy as np
from scipy import ndimage, signal, stats



root = Tk()
root.title("Root_Window")
root.geometry("800x800")



C:/Users/james.harkin/Documents
class main:

    #root path to search for images
    root_path = "C:/Users/James/Documents/EAUSQA"
     
    class initialise_analysis:
        
        def __init__(self, root_path):
            self.images_directory = self.Select_Directory(root_path, 
                "Select directory containing all dicom images to be sorted.")
            self.dcm_paths = self.Get_Files(self.images_directory)
            self.dcm_dict = self.Get_Dicom_Dict(self.dcm_paths)
            self.Get_Acquisition_Details()      
            
            
        def Select_Directory(self, initial_dir, window_title):
            """
            Parameters
            ----------
            initial_dir: str
                Path at which the file dialogue opens
            window_title: str
                Title of the filedialogue window.
                
            Returns
            -------
            user_selected_directory : str
            
            Uses tkinter filedialog.  Asks the user to select a directory and
            returns the user_selected_directory.
            """
            directory_window = Toplevel()
            directory_window.wm_attributes('-topmost', 1)
            directory_window.withdraw()
            user_selected_directory =  filedialog.askdirectory(parent=directory_window, 
                                                               initialdir=initial_dir, 
                                                               title=window_title)
            directory_window.destroy()
            return user_selected_directory
    
        def Get_Files(self, starting_dir):
            """
            Parameters
            ----------
            starting_dir: str
                Every file with no extension and not called VERSION inside 
                starting_dir will be assumed to be a dicom of interest.
            
            Returns
            -------
            paths : list of str
                list of dicom_paths inside starting_dir.
            """
            paths = []
             
            for dirName, subdirList, fileList in os.walk(starting_dir):
                #individual channel data
                for filename in fileList:
                    if filename != "VERSION" and filename[-5:] != ".xlsx":
                        paths.append(os.path.join(dirName,filename))
                    
            return paths
        
        def Get_Dicom_Dict(self, dicom_paths):
            """
            Parameters
            ----------
            dicom_paths : list of strings
                List of all dicom paths
            Returns
            -------
            dcm_dict : dict
                Dictionary of dicoms of the form {<path>:<dicom>}
            """
            dcm_dict = {}
            for dicom_path in dicom_paths:
                dcm_dict[dicom_path] = {}
                dcm = pydicom.read_file(dicom_path)
                dcm_dict[dicom_path]["Dcm"] = dcm
            return dcm_dict
        
        def Get_Acquisition_Details(self):
            analysis_complete = IntVar()               
                    
            while analysis_complete.get() != True:
                top = Toplevel()
                top.title("Confirm Analysis To Perform")
                top.geometry('400x150')
                top.wm_attributes('-topmost', 1)
                same_FOV = IntVar()
                ask_parameters = IntVar()
                gt = IntVar()
                hc = IntVar()
                lc = IntVar()
                Checkbutton(top, text="Same_FOV?", variable=same_FOV ).grid(row=0, column=0, sticky=W)
                Checkbutton(top, text="Ask Acquisition Parameters?", variable=ask_parameters).grid(row=1, column=0, sticky=W)
                Checkbutton(top, text="Grey Targets", variable=gt).grid(row=2, column=0, sticky=W)
                Checkbutton(top, text="High Contrast Targets", variable=hc).grid(row=3, column=0, sticky=W)
                Checkbutton(top, text="Low Contrast Targets", variable=lc).grid(row=4, column=0, sticky=W)
                Button(top, text="Perform Analysis", command=lambda:  analysis_complete.set(True)).grid(row=5, column=0,sticky=W)
                top.wait_variable(analysis_complete)                   
                top.destroy()
            self.same_FOV = same_FOV.get()
            self.ask_parameters = ask_parameters
            self.gt = gt.get()            
            self.hc = hc.get()
            self.lc = lc.get()
            
            
            
    class analyse_dicoms:
        class define_imageparameters:
            bore_diam_mm = 18.4
            gt_diam_mm = 10
            target_names = ["-9.3dB","-5.7dB", "0dB","7dB","9.7dB"]
            def __init__(self, dcm, get_acq_param=True, pixel_size=None, image_center=None):
                
                self.img = dcm.pixel_array
                self.grey = cv2.cvtColor(self.img, cv2.COLOR_BGR2GRAY)
                self.displayimg = np.array(Image.fromarray(self.img))          
                
                
                if get_acq_param == True:
                    self.Get_Acquisition_Parameters(dcm)
                
                self.pixel_size = pixel_size
                self.image_center = image_center
                    
                
                
            def click_event(self, event, x, y, flags, params):
      
                # checking for left mouse clicks
                if event == cv2.EVENT_LBUTTONDOWN:
              
                    # displaying the coordinates
                    # on the Shell
                    self.L_d = [x,y]
                    print(self.L_d)
                
                if event == cv2.EVENT_LBUTTONUP:
              
                    # displaying the coordinates
                    # on the Shell
                    self.L_u = [x,y]
              
            def circle(self, event, x, y, flags, window_name):
                
                # checking for left mouse clicks
                 
                if event == cv2.EVENT_MOUSEMOVE:
                    self.displayimg= copy.deepcopy(self.img_copy)   
                    cv2.imshow(window_name, self.displayimg)
                    try:
                        cv2.circle(self.displayimg,self.L_d,self.circle_radius_px ,(255,0,0),3)
                    except SystemError:
                        pass
                     
                if event == cv2.EVENT_RBUTTONDOWN:
                    self.img_copy = copy.deepcopy(self.displayimg)
                    cv2.circle(self.displayimg,(x,y),self.circle_radius_px ,(255,0,0),3) 
                
                if event == cv2.EVENT_LBUTTONDOWN:
                    self.L_d = (x,y)
                    
            def target_event(self, event, x, y, flags, window_name):
                if event == cv2.EVENT_LBUTTONDOWN:
                    # displaying the coordinates
                    # on the Shell
                    self.L_d = (x,y)
                    cv2.circle(self.displayimg,self.L_d,self.circle_radius_px,(0,0,255),1) 
                    self.targets.append(self.L_d)
                    
                if event == cv2.EVENT_RBUTTONDOWN:
                    self.displayimg = copy.deepcopy(self.img_copy)
                    cv2.imshow(window_name,self.displayimg)
                    self.targets = []
              
            def Get_Acquisition_Parameters(self,dcm):
                window_name ='Confirm Acquisition Parameters'
                cv2.namedWindow(window_name)
                while True:
                    cv2.imshow(window_name,self.displayimg)
                    key = cv2.waitKey(1) & 0xFF
                    # if the 'c' key is pressed, break from the loop
                    top = Toplevel()
                    top.title("Confirm Acquisition Parameters")
                    top.geometry('300x150')
                    top.wm_attributes('-topmost', 1)
                    
                    Scanner_Name = StringVar()
                    Scanner_ID = StringVar()
                    gain = IntVar()
                    frequency = IntVar()
                    FOV_mm = IntVar()
                    dynamic_range = IntVar()
                    proceed = IntVar()
                    
                    Scanner_Name.set(dcm.PatientID)
                    Scanner_ID.set(dcm.PatientID)
                    gain.set(50)
                    FOV_mm.set(110)
                    frequency.set(13)
                    dynamic_range.set(65)
                    
                    Label(top, text = "Scanner Name").grid(row=0, column=0, sticky=W)
                    Entry(top, width = 15, textvariable= Scanner_Name).grid(row=0, column=1, sticky=W)
                    Label(top, text = "Scanner ID").grid(row=1, column=0, sticky=W)
                    Entry(top, width = 15, textvariable= Scanner_ID).grid(row=1, column=1, sticky=W)
                    Label(top, text = "Gain (dB)").grid(row=2, column=0, sticky=W)
                    Entry(top, width = 15, textvariable=gain).grid(row=2, column=1, sticky=W)
                    Label(top, text = "Frequency (MHz)").grid(row=3, column=0, sticky=W)
                    Entry(top, width = 15, textvariable=frequency).grid(row=3, column=1, sticky=W)
                    Label(top, text = "FOV (mm)").grid(row=4, column=0, sticky=W)
                    Entry(top, width = 15, textvariable=FOV_mm).grid(row=4, column=1, sticky=W)
                    Label(top, text = "Dynamic Range (dB").grid(row=5, column=0, sticky=W)
                    Entry(top, width = 15, textvariable=dynamic_range).grid(row=5, column=1, sticky=W)
                    Button(top, text="Proceed", command=lambda:  proceed.set(True)).grid(row=6, column=0,sticky=W)
                    top.wait_variable(proceed)                   
                    top.destroy()
                    
                    if proceed.get() == True:
                        break
                cv2.destroyAllWindows()
                self.scanner_name = Scanner_Name.get()
                self.scanner_ID = Scanner_ID.get()            
                self.gain = gain.get()
                self.frequency = frequency.get()
                self.FOV_mm = FOV_mm.get()
                self.DR = dynamic_range.get()
            
            
            def Get_Size(self):
                window_name ='Click two points 5cm apart, c to confirm selection'
                cv2.namedWindow(window_name)
                cv2.setMouseCallback(window_name, self.click_event)
                # keep looping until the 'q' key is pressed
                while True:
                    
                    cv2.imshow(window_name,self.displayimg)
                    key = cv2.waitKey(1) & 0xFF
                    # if the 'c' key is pressed, break from the loop
                    if key == ord("c"):
                        break
                cv2.destroyAllWindows()                
                pixel_distance = ((self.L_u[0]-self.L_d[0])**2 + (self.L_u[1]-self.L_d[1])**2)**0.5
                pixel_size_mm = 50/pixel_distance
                print(pixel_size_mm)
                return pixel_size_mm
                
            def Get_Center(self):
                self.circle_radius_px = int(self.bore_diam_mm/(2*self.pixel_size))
                self.img_copy= copy.deepcopy(self.displayimg)
                self.L_d = NONE
                window_name = 'left click to select bore center, right click to display bore size, c to confirm selection'
                cv2.namedWindow(window_name)
                # keep looping until the 'q' key is pressed
                cv2.setMouseCallback(window_name, self.circle, window_name)
                    
                while True:
                    cv2.imshow(window_name,self.displayimg)
                    key = cv2.waitKey(1) & 0xFF
                    # if the 'c' key is pressed, break from the loop
                    if key == ord("c"):
                        break
                    
                    
                cv2.destroyAllWindows()                
                return self.L_d
            
            
            def Select_Grey_Targets(self):
                target_centers = {}
                for target_name in self.target_names:
                    self.circle_radius_px  = int(self.gt_diam_mm/(2*self.pixel_size))
                    self.img_copy= copy.deepcopy(self.displayimg)
                    self.L_d = NONE
                    window_name = 'left click to select' + target_name + 'target center, right click to display target size, c to confirm selection'
                    cv2.namedWindow(window_name)
                    # keep looping until the 'q' key is pressed
                    cv2.setMouseCallback(window_name, self.circle, window_name)
                        
                    while True:
                        cv2.imshow(window_name,self.displayimg)
                        key = cv2.waitKey(1) & 0xFF
                        # if the 'c' key is pressed, break from the loop
                        if key == ord("c"):
                            break
                    cv2.destroyAllWindows()   
                    target_centers[target_name] = self.L_d             
                return target_centers
                
            def Get_Targets(self, target_name, target_size=5):
                self.circle_radius_px = int(target_size/(2*self.pixel_size))
                self.img_copy= copy.deepcopy(self.displayimg)
                self.L_d = NONE
                self.targets = []
                window_name = target_name + 'left click to select high contrast target, right click to cancel selection c to confirm selection'
                cv2.namedWindow(window_name)
                # keep looping until the 'q' key is pressed
                cv2.setMouseCallback(window_name, self.target_event, window_name)
                    
                while True:
                    cv2.imshow(window_name,self.displayimg)
                    key = cv2.waitKey(1) & 0xFF
                    # if the 'c' key is pressed, break from the loop
                    if key == ord("c"):
                        break
                cv2.destroyAllWindows()               
                return self.targets
            
            def Find_Target_Center(self, approx_centers, search_size_mm=5, k_size = (1,1), check_centers=True):
                search_size_px = int(search_size_mm/(2*self.pixel_size))
                blurred_img = cv2.GaussianBlur(self.grey, k_size, 0)
                calculated_centers = []
                for approx_center in approx_centers:
                    x_0 = approx_center[0]-search_size_px
                    x_1 = approx_center[0]+search_size_px
                    y_0 = approx_center[1]-search_size_px
                    y_1 = approx_center[1]+search_size_px
                    search_image = blurred_img[y_0:y_1,x_0:x_1]
                    (minVal, maxVal, minLoc, maxLoc) = cv2.minMaxLoc(search_image)
                    calculated_centers.append((x_0+maxLoc[0],y_0+maxLoc[1]))
                    
                if check_centers == True:
                    for center in calculated_centers:
                        cv2.drawMarker(self.displayimg, center,(255,0,0))
                    while True:
                         cv2.imshow("centers",self.displayimg)
                         key = cv2.waitKey(1) & 0xFF
                         # if the 'c' key is pressed, break from the loop
                         if key == ord("c"):
                             break
                    cv2.destroyAllWindows()               
                return calculated_centers
                    
                
            
        class calculate_results:
            bore_diam_mm = 18.4
            gt_diam_mm = 10
            gt_attenuations = [-9.3,-5.7,0,7,9.7]
            target_names = ["-9.3dB","-5.7dB", "0dB","7dB","9.7dB"]
            def __init__(self, dcm, pixel_size, image_center):
                self.dcm_np = dcm.pixel_array.astype(float)
                self.h, self.w = self.dcm_np.shape[:2]
                self.pixel_size = pixel_size
                self.image_center = image_center

            
            def create_circular_mask(self, h, w, center, radius):
                
                Y, X = np.ogrid[:h, :w]
                dist_from_center = np.sqrt((X - center[0])**2 + (Y-center[1])**2)
            
                mask = dist_from_center <= radius
                return mask 
            
            def create_corona_mask(self, h, w, center, inner_radius, outer_radius):
                #creates coronal mask of same area as circle encompased
                inner_mask = self.create_circular_mask(h, w, center, inner_radius)
                outer_mask = self.create_circular_mask(h, w, center, outer_radius)
                corona_mask = np.bitwise_xor(outer_mask, inner_mask)
                return corona_mask
            
            def Calculate_Penetration_Depth(self, background_center, required_int, target_radius):
                
                x_i = background_center[0] - self.image_center[0]
                y_i = self.image_center[1] - background_center[1]
                z = complex(x_i, y_i)
                pen_angle = math.degrees(cmath.phase(z))
                penetration_depth = None
                depths = []
                intensities = []
                #start at 15mm depth
                depth = abs(15/self.pixel_size)
                
                while depth < 412:
                    a = depth*math.cos(pen_angle)
                    b = depth*math.sin(pen_angle)       
                    measurement_x = self.image_center[1]+depth*math.cos(pen_angle)
                    measurement_y = self.image_center[0]-depth*math.sin(pen_angle)
                    measurement_mask = self.create_circular_mask(self.h, self.w, [measurement_y,measurement_x], target_radius)
                    masked_np = self.dcm_np.copy()
                    masked_np[~measurement_mask] = np.NaN
                    average_intensity = np.nanmean(masked_np)
                    depths.append(depth)
                    intensities.append(average_intensity)
                    
                    if average_intensity <= required_int and penetration_depth == None:
                        penetration_depth = depth
                    depth = depth + 1
                    
                try:
                    penetration_depth_mm = penetration_depth*self.pixel_size
                except TypeError:
                    penetration_depth_mm = 100
                
                return penetration_depth_mm
                
                
                
                
                
                
                
            def GT_Contrast_Calculate(self, target_centers, penetration_calc=True):
                # only look at central region of target
                target_radius = int(self.gt_diam_mm*0.8/(2*self.pixel_size))
                self.gt_average_intensity = {}
                self.gt_variance = {}
                self.gt_SNR = {}
                self.gt_cnr = {}
                
                for target_name in self.target_names:
                    target_center = target_centers[target_name]
                    target_mask = self.create_circular_mask(self.h, self.w, target_center, target_radius)
                    masked_np = self.dcm_np.copy()
                    masked_np[~target_mask] = np.NaN
                    heading_average_int = target_name + " average"
                    heading_var = target_name + " var"
                    self.gt_average_intensity[heading_average_int] = np.nanmean(masked_np)
                    self.gt_variance[heading_var] = np.nanvar(masked_np)
                    
                self.backgroundint = self.gt_average_intensity["0dB average"] 
                for target_name in self.target_names:
                    heading_CNR = target_name +" CNR"
                    heading_SNR = target_name +" SNR"
                    self.gt_cnr[heading_CNR] = abs(self.gt_average_intensity[target_name+" average"]-self.backgroundint)/((self.gt_variance[target_name+" var"]+self.gt_variance["0dB var"])**0.5)
                    self.gt_SNR[heading_SNR] = self.gt_average_intensity[target_name+" average"]/(self.gt_variance["0dB var"]**0.5)
                
                if penetration_calc==True:
                    average_intensities = list(self.gt_average_intensity.values())
                    slope, intercept, r_value, p_value, std_err = stats.linregress(self.gt_attenuations,average_intensities)
                    if r_value<0.8:
                        print("could not establish -6dB  Penetration depth not performed")
                        pen = None
                    else:
                        required_int = (-6*slope) + intercept
                        if required_int < 0:
                            required_int = 0
                        pen = self.Calculate_Penetration_Depth(target_centers["0dB"],required_int,target_radius)
                
                return self.gt_cnr, self.gt_SNR, self.gt_average_intensity, self.gt_variance, pen
                
                
                
            def Calculate_Distances(self, targets_1, targets_2):
                no_targets = len(targets_1)
                self.arc_distances_mm = []
                self.radial_distances_1_mm = []
                self.radial_distances_2_mm = []
                for i in range(no_targets):
                    targets_1_center_1 = targets_1[i]
                    targets_2_center_1 = targets_2[i]
                    self.arc_distances_mm.append(self.pixel_size*((targets_1_center_1[0]-targets_2_center_1[0])**2 + (targets_1_center_1[1]-targets_2_center_1[1])**2)**0.5)
                    
                    try:
                        targets_1_center_2 = targets_1[i+1]
                        self.radial_distances_1_mm.append(self.pixel_size*((targets_1_center_1[0]-targets_1_center_2[0])**2 + (targets_1_center_1[1]-targets_1_center_2[1])**2)**0.5)
                        targets_2_center_2 = targets_2[i+1]
                        self.radial_distances_2_mm.append(self.pixel_size*((targets_2_center_1[0]-targets_2_center_2[0])**2 + (targets_2_center_1[1]-targets_2_center_2[1])**2)**0.5)
                    except IndexError:
                        pass
                return self.radial_distances_1_mm, self.radial_distances_2_mm, self.arc_distances_mm
                
                
            def Calculate_Resolution(self, targets, search_size_mm):
                search_size_px = int(search_size_mm/(2*self.pixel_size))
                self.radial_target_sizes_mm = []
                self.tang_target_sizes_mm = []
                self.radial_MIP_target_sizes_mm = []
                self.tang_MIP_target_sizes_mm = []
                for target_center in targets:
                    x_0 = target_center[0]-search_size_px
                    x_1 = target_center[0]+search_size_px+1
                    y_0 = target_center[1]-search_size_px
                    y_1 = target_center[1]+search_size_px+1
                    x_i = target_center[0] - self.image_center[0]
                    y_i = self.image_center[1] - target_center[1]
                    rot_angle = math.degrees(math.atan((self.image_center[0]-target_center[0])/(target_center[1]-self.image_center[1])))
                    #print(rot_angle)
                    image_roi = self.dcm_np[y_0:y_1,x_0:x_1]
                    image_roi_rot = ndimage.rotate(image_roi,rot_angle,reshape=False)
                    central_pixel_int = image_roi_rot[search_size_px,search_size_px]
                    
                    radial_slice = image_roi_rot[:,search_size_px,0]
                    radial_MIP = np.max(image_roi_rot[:,:,0],axis=1)
                    rad_pk = np.argmax(image_roi_rot[:,search_size_px,0])
                    tang_slice = image_roi_rot[search_size_px,:,0]
                    #cruvaturae of phantom Maximum intensity projection user for tangental resolution
                    tang_MIP = np.max(image_roi_rot[:,:,0],axis=0)
                    tang_pk = np.argmax(image_roi_rot[search_size_px,:,0])
                    
                    #NOTFWHM  Promenence
                    radial_peak_properties = signal.peak_widths(radial_slice, np.array([rad_pk]), rel_height=0.5)
                    tang_peak_properties = signal.peak_widths(tang_slice, np.array([tang_pk]), rel_height=0.5)
                    radial_MIP_peak_properties = signal.peak_widths(radial_MIP, np.array([rad_pk]), rel_height=0.5)
                    tang_MIP_peak_properties = signal.peak_widths(tang_MIP, np.array([tang_pk]), rel_height=0.5)
                    
                    
                    
                    
                    self.radial_target_sizes_mm.append(self.pixel_size*radial_peak_properties[0][0])
                    self.tang_target_sizes_mm.append(self.pixel_size*tang_peak_properties[0][0])
                    self.radial_MIP_target_sizes_mm.append(self.pixel_size*radial_MIP_peak_properties[0][0])
                    self.tang_MIP_target_sizes_mm.append(self.pixel_size*tang_MIP_peak_properties[0][0])
                
                return self.radial_target_sizes_mm, self.tang_target_sizes_mm, self.radial_MIP_target_sizes_mm, self.tang_MIP_target_sizes_mm
                
                
            def Calculate_Low_Contrast_Object_Detectibility(self, targets, LCO_diam_mm):
                # only look at central region of target
                target_radius = int(LCO_diam_mm*0.8/(2*self.pixel_size))
                target_area = target_radius**2
                #corona same area as target
                internal_radius = int(LCO_diam_mm*1.2/(2*self.pixel_size))
                external_radius = int((target_area + internal_radius**2)**0.5)
                
                self.LCO_average_intensity = []
                self.LCO_variance = []
                self.LCO_corona_average_intensity = []
                self.LCO_corona_variance = []
                
                self.LCO_contrast = []
                self.LCO_CNR = []
                for target_center in targets:
                    target_mask = self.create_circular_mask(self.h, self.w, target_center, target_radius)
                    corona_mask = self.create_corona_mask(self.h, self.w, target_center, internal_radius, external_radius)
                    
                    
                    
                    LCO_masked_np = self.dcm_np.copy()
                    LCO_masked_np[~target_mask] = np.NaN
                    self.LCO_average_intensity.append(np.nanmean(LCO_masked_np))
                    self.LCO_variance.append(np.nanvar(LCO_masked_np))
                    
                    LCO_corona_masked_np = self.dcm_np.copy()
                    LCO_corona_masked_np[~corona_mask] = np.NaN
                    self.LCO_corona_average_intensity.append(np.nanmean(LCO_corona_masked_np))
                    self.LCO_corona_variance.append(np.nanvar(LCO_corona_masked_np))
                for i in range(len(targets)):
                    self.LCO_CNR.append(abs(self.LCO_corona_average_intensity[i]-self.LCO_average_intensity[i])/((self.LCO_corona_variance[i]+self.LCO_variance[i])**0.5))
                    self.LCO_contrast.append(abs(self.LCO_corona_average_intensity[i]-self.LCO_average_intensity[i]))
                
                return self.LCO_CNR, self.LCO_contrast
        
    
        def __init__(self, initialised, THD=False):      
            self.params_dict = {}
            self.results = {}
            self.pixel_size = None
            self.image_center = None
            self.grey_target_centers = None
            self.hc_centers_1 = None
            self.hc_centers_2 = None
            self.lc_centers_4mm = None
            self.lc_centers_2mm = None
            self.lc_centers_1mm = None
            
            same_location = initialised.same_FOV
            ask_parameters = initialised.ask_parameters
            
            for dcm_path in initialised.dcm_dict:
                analysis_complete = False
                self.params_dict[dcm_path] = {}
                self.results[dcm_path] = {}
                dcm = initialised.dcm_dict[dcm_path]["Dcm"]
                

                self.results[dcm_path]["File_Name"] = ntpath.basename(dcm_path)
                if ask_parameters == True:   
                    self.basic_params = self.define_imageparameters(dcm)
                    self.results[dcm_path]["Scanner_Name"] = self.basic_params.scanner_name
                    self.results[dcm_path]["Scanner_ID"] = self.basic_params.scanner_ID
                    self.results[dcm_path]["Gain"] = self.basic_params.gain
                    self.results[dcm_path]["Frequency"] = self.basic_params.frequency
                    self.results[dcm_path]["FOV"] = self.basic_params.FOV_mm
                    self.results[dcm_path]["Dynamic_Range"] = self.basic_params.DR
                
                else:
                    self.results[dcm_path]["Scanner_Name"] = "Not_added"
                    self.results[dcm_path]["Scanner_ID"] = "Not_added"
                    self.results[dcm_path]["Gain"] = "Not_added"
                    self.results[dcm_path]["Frequency"] = "Not_added"
                    self.results[dcm_path]["FOV"] = "Not_added"
                    self.results[dcm_path]["Dynamic_Range"] = "Not_added"
                
                
                if self.pixel_size!=None and self.image_center!=None and same_location ==True:
                    pass
                elif THD==True:
                    #specific application for machine with no displayed scale/ DICOM header
                    self.pixel_size = 0.17094
                    self.image_center = self.define_imageparameters(dcm,False, self.pixel_size).Get_Center()
                else:
                    self.pixel_size = self.define_imageparameters(dcm,False).Get_Size()
                    self.image_center = self.define_imageparameters(dcm,False, self.pixel_size).Get_Center()
                self.params_dict[dcm_path]["pixel_size"] = self.pixel_size
                self.params_dict[dcm_path]["image_center"] = self.image_center
                
                
                if initialised.gt == True:
                    if self.grey_target_centers!=None and same_location==True:
                        pass
                    else:
                        self.grey_target_centers = self.define_imageparameters(dcm,False, self.pixel_size,self.image_center).Select_Grey_Targets()
                    self.params_dict[dcm_path]["gt_centers"] = self.grey_target_centers
                    self.results[dcm_path]["gt_CNR"], self.results[dcm_path]["gt_SNR"], self.results[dcm_path]["gt_average_intensity"], self.results[dcm_path]["gt_variance"], self.results[dcm_path]["gt_pen"] = self.calculate_results(dcm,self.pixel_size,self.image_center).GT_Contrast_Calculate(self.grey_target_centers)
                if initialised.hc == True:            
                    if self.hc_centers_1!=None and same_location==True and self.hc_centers_2!=None:
                        self.hc_centers_1 = self.define_imageparameters(dcm,False,self.pixel_size,self.image_center).Find_Target_Center(high_contrast_centers_1_manual, check_centers=False)
                        self.hc_centers_2 = self.define_imageparameters(dcm,False,self.pixel_size,self.image_center).Find_Target_Center(high_contrast_centers_2_manual, check_centers=False)
                    else:
                        high_contrast_centers_1_manual = self.define_imageparameters(dcm,False,self.pixel_size,self.image_center).Get_Targets("high_contrast_centers_1")
                        high_contrast_centers_1_calculated = self.define_imageparameters(dcm,False,self.pixel_size,self.image_center).Find_Target_Center(high_contrast_centers_1_manual, check_centers=False)
                        high_contrast_centers_2_manual = self.define_imageparameters(dcm,False,self.pixel_size,self.image_center).Get_Targets("high_contrast_centers_2")
                        high_contrast_centers_2_calculated = self.define_imageparameters(dcm,False,self.pixel_size,self.image_center).Find_Target_Center(high_contrast_centers_2_manual, check_centers=False)
                    
                        
                        self.hc_centers_1 = high_contrast_centers_1_calculated
                        self.hc_centers_2 = high_contrast_centers_2_calculated
                        
                    self.params_dict[dcm_path]["hc_centers_1"] = self.hc_centers_1
                    self.params_dict[dcm_path]["hc_centers_2"] = self.hc_centers_2
                    
                    self.results[dcm_path]["radial_distances_1"], self.results[dcm_path]["radial_distances_2"], self.results[dcm_path]["arc_distances"] = self.calculate_results(dcm,self.pixel_size,self.image_center).Calculate_Distances(self.hc_centers_1, self.hc_centers_2)
                    self.results[dcm_path]["radial_target_sizes_1"], self.results[dcm_path]["tang_target_sizes_1"], self.results[dcm_path]["radial_MIP_target_sizes_1"], self.results[dcm_path]["tang_MIP_target_sizes_1"] = self.calculate_results(dcm,self.pixel_size,self.image_center).Calculate_Resolution(self.hc_centers_1, 5)
                    self.results[dcm_path]["radial_target_sizes_2"], self.results[dcm_path]["tang_target_sizes_2"], self.results[dcm_path]["radial_MIP_target_sizes_2"], self.results[dcm_path]["tang_MIP_target_sizes_2"] = self.calculate_results(dcm,self.pixel_size,self.image_center).Calculate_Resolution(self.hc_centers_2, 5)
                                
                if initialised.lc == True:
                    if self.lc_centers_4mm!=None and same_location==True:
                        pass
                    else:
                        self.lc_centers_4mm = self.define_imageparameters(dcm,False,self.pixel_size,self.image_center).Get_Targets("low_contrast_centers_4mm", 4)
                    #self.lc_centers_4mm = self.define_imageparameters(dcm,False,self.pixel_size,self.image_center).Get_Targets("low_contrast_centers_4mm", 4)
                    self.params_dict[dcm_path]["low_contrast_centers_4mm"] = self.lc_centers_4mm
                    self.results[dcm_path]["LCO_4mm_CNR"], self.results[dcm_path]["LCO_4mm_contrast"] = self.calculate_results(dcm,self.pixel_size,self.image_center).Calculate_Low_Contrast_Object_Detectibility(self.lc_centers_4mm, 4)
                    
                    if self.lc_centers_2mm!=None and same_location==True:
                        pass
                    else:
                        self.lc_centers_2mm = self.define_imageparameters(dcm,False,self.pixel_size,self.image_center).Get_Targets("low_contrast_centers_2mm", 2)
                    #self.lc_centers_2mm = self.define_imageparameters(dcm,False,self.pixel_size,self.image_center).Get_Targets("low_contrast_centers_2mm", 2)
                    self.params_dict[dcm_path]["low_contrast_centers_2mm"] = self.lc_centers_2mm
                    self.results[dcm_path]["LCO_2mm_CNR"], self.results[dcm_path]["LCO_2mm_contrast"] = self.calculate_results(dcm,self.pixel_size,self.image_center).Calculate_Low_Contrast_Object_Detectibility(self.lc_centers_2mm, 2)
                    
                    if self.lc_centers_1mm!=None and same_location==True:
                        pass
                    else:
                        self.lc_centers_1mm = self.define_imageparameters(dcm,False,self.pixel_size,self.image_center).Get_Targets("low_contrast_centers_1mm", 1)
                    #self.lc_centers_1mm = self.define_imageparameters(dcm,False,self.pixel_size,self.image_center).Get_Targets("low_contrast_centers_1mm",1)
                    self.params_dict[dcm_path]["low_contrast_centers_1mm"] = self.lc_centers_1mm
                    self.results[dcm_path]["LCO_1mm_CNR"], self.results[dcm_path]["LCO_1mm_contrast"] = self.calculate_results(dcm,self.pixel_size,self.image_center).Calculate_Low_Contrast_Object_Detectibility(self.lc_centers_1mm, 1)
                            
                
                
    
    class export_to_excel:
        """
        A class used to exports  results to excel
        
        Attributes
        ----------
        wb : Excel Workbook
        
        ws : Excel Worksheet

        """
        headings_dict = {"File_Name":["File Name"],"Scanner_Name":["Scanner Name"],"Scanner_ID":["Scanner ID"],"Gain": ["Gain"], 
                         "Frequency":["Frequency"],"FOV":["FOV"],"Dynamic_Range":["Dynamic Range"],
                         "gt_CNR":["-9.3dB CNR","-5.7dB CNR", "0dB CNR","7dB CNR","9.7dB CNR"],
                         "gt_SNR":["-9.3dB SNR","-5.7dB SNR", "0dB SNR","7dB SNR","9.7dB SNR"],
                         "gt_average_intensity":["-9.3dB average","-5.7dB average", "0dB average","7dB average","9.7dB average"],
                         "gt_variance":["-9.3dB var","-5.7dB var", "0dB var","7dB var","9.7dB var"],
                         "gt_pen":["Penetration Depth"],
                         "radial_distances_1":["raddist_1_1","raddist_1_2","raddist_1_3","raddist_1_4","raddist_1_5","raddist_1_6"],
                         "radial_distances_2":["raddist_2_1","raddist_2_2","raddist_2_3","raddist_2_4","raddist_2_5","raddist_2_6"],
                         "arc_distances":["arcdist_1","arcdist_2","arcdist_3","arcdist_4","arcdist_5","arcdist_6","arcdist_7"],
                         "radial_target_sizes_1":["radsize_1","radsize_1_2","radsize_1_3","radsize_1_4","radsize_1_5","radsize_1_6","radsize_1_7"],
                         "radial_target_sizes_2":["radsize_2_1","radsize_2_2","radsize_2_3","radsize_2_4","radsize_2_5","radsize_2_6","radsize_2_7"],
                         "radial_MIP_target_sizes_1":["MIP_radsize_1","MIP_radsize_1_2","MIP_radsize_1_3","MIP_radsize_1_4","MIP_radsize_1_5","MIP_radsize_1_6","MIP_radsize_1_7"],
                         "radial_MIP_target_sizes_2":["MIP_radsize_2_1","MIP_radsize_2_2","MIP_radsize_2_3","MIP_radsize_2_4","MIP_radsize_2_5","MIP_radsize_2_6","MIP_radsize_2_7"],
                         "tang_target_sizes_1":["tangsize_1_1","tangsize_1_2","tangsize_1_3","tangsize_1_4","tangsize_1_5","tangsize_1_6","tangsize_1_7"],
                         "tang_target_sizes_2":["tangsize_2_1","tangsize_2_2","tangsize_2_3","tangsize_2_4","tangsize_2_5","tangsize_2_6","tangsize_2_7"],
                         "tang_MIP_target_sizes_1":["MIP_tangsize_1_1","MIP_tangsize_1_2","MIP_tangsize_1_3","MIP_tangsize_1_4","MIP_tangsize_1_5","MIP_tangsize_1_6","MIP_tangsize_1_7"],
                         "tang_MIP_target_sizes_2":["MIP_tangsize_2_1","MIP_tangsize_2_2","MIP_tangsize_2_3","MIP_tangsize_2_4","MIP_tangsize_2_5","MIP_tangsize_2_6","MIP_tangsize_2_7"],
                         "LCO_4mm_CNR":["LCO_CNR_4mm_1","LCO_CNR_4mm_2","LCO_CNR_4mm_3","LCO_CNR_4mm_4"], 
                         "LCO_4mm_contrast":["LCO_contrast_4mm_1","LCO_contrast_4mm_2","LCO_contrast_4mm_3","LCO_contrast_4mm_4"],
                         "LCO_2mm_CNR":["LCO_CNR_2mm_1","LCO_CNR_2mm_2","LCO_CNR_2mm_3","LCO_CNR_2mm_4"],
                         "LCO_2mm_contrast":["LCO_contrast_2mm_1","LCO_contrast_2mm_2","LCO_contrast_2mm_3","LCO_contrast_2mm_4"],
                         "LCO_1mm_CNR":["LCO_CNR_1mm_1","LCO_CNR_1mm_2","LCO_CNR_1mm_3","LCO_CNR_1mm_4"],
                         "LCO_1mm_contrast":["LCO_contrast_1mm_1","LCO_contrast_1mm_2","LCO_contrast_1mm_3","LCO_contrast_1mm_4"]}
        
        def __init__(self, results, archive_dir, file_name="eaus_results.xlsx"):
            
            self.Export_Data(archive_dir, file_name, results)
            self.wb.close()
        
        
        def Open_Spreadsheet(self, directory, file_name):
            
            if not os.path.exists(directory):
                    os.makedirs(directory)
            os.chdir(directory)
            
            """Open correct spreadsheet. If it doesn't exist create it"""
            if os.path.isfile(os.path.join(directory, file_name)):
                self.wb = openpyxl.load_workbook(file_name)
            else:
                #first SNR data for this scanner; create new file
                self.wb = openpyxl.Workbook()
                self.wb.save(file_name)
                print("New Spreadsheet created.")
            

        def Initialise_Sheet(self, ws_name):
            
            self.wb.create_sheet(ws_name)
            self.ws = self.wb[ws_name]
            i=0
            for category in self.headings_dict:
                for heading in self.headings_dict[category]:
                    #write <headings> in the first row of the worksheet
                    i=i+1
                    self.ws.cell(column=i, row=1, value=heading)
                    

        def Open_Sheet(self, file_name, sheet_name):
            
            if not sheet_name in self.wb.sheetnames:
                self.Initialise_Sheet(sheet_name)
                self.wb.save(file_name)
                print("New Sheet created.  This should be the first SNR measurement for this coil")
            else:
                self.ws = self.wb[sheet_name]


        def Input_Results(self, data_for_input):
            new_row = self.ws.max_row+1
            i=0
            for category in self.headings_dict:
                j=0
                for heading in self.headings_dict[category]:
                    i=i+1
                    try:
                        self.ws.cell(column=i, row=new_row, value=data_for_input[category])
                    except KeyError:
                        pass
                    except ValueError:
                        try:
                            self.ws.cell(column=i, row=new_row, value=data_for_input[category][heading])
                        except TypeError:
                            try:
                                self.ws.cell(column=i, row=new_row, value=data_for_input[category][j])
                            except IndexError:
                                pass
                    j=j+1
                    
        def Export_Data(self, directory, file_name, results):   
            
            self.Open_Spreadsheet(directory, file_name)  
            #Initialise sheets if not already created with the correct headings.  
            #Headings are specified in the SNR_data_dict[sheet_name] dictionary.
            for analysis in results:                
                sheet_name = results[analysis]["Scanner_ID"]
                self.Open_Sheet(file_name, sheet_name)    
                
                self.Input_Results(results[analysis])
            self.wb.save(file_name)
            self.wb.close()
    
    def __init__(self, master):
        """
        Parameters
        ----------
        master : tkinter root
        """
        master.withdraw()
        initialised = self.initialise_analysis(self.root_path)
        completed_analysis = self.analyse_dicoms(initialised)
        self.export_to_excel(completed_analysis.results, initialised.images_directory)
        master.destroy()
        
        
main(root)

root.mainloop()

